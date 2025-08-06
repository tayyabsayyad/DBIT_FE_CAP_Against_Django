from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import views as auth_views
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse
from openpyxl import Workbook
from .models import FirstYearAdmission, SecondYearAdmission
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from .models import FirstYearAdmission, SecondYearAdmission
from django.template.loader import render_to_string
from weasyprint import HTML
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .forms import FirstYearAdmissionForm, SecondYearAdmissionForm, RegistrationForm
from .models import FirstYearAdmission, SecondYearAdmission


@login_required
def dashboard(request):
    return render(request, 'dashboard.html')

@login_required
def first_year_admission_view(request):
    try:
        existing_app = request.user.first_year_application
        return redirect('admissionapp:first_year_admission_detail')
    except FirstYearAdmission.DoesNotExist:
        existing_app = None

    if request.method == 'POST':
        form = FirstYearAdmissionForm(request.POST)
        if form.is_valid():
            admission = form.save(commit=False)
            admission.user = request.user
            admission.save()
            return redirect('admissionapp:first_year_admission_detail')
    else:
        form = FirstYearAdmissionForm()

    return render(request, 'first_year_form.html', {'form': form})

@login_required
def second_year_admission_view(request):
    try:
        existing_app = request.user.second_year_application
        return redirect('admissionapp:second_year_admission_detail')
    except SecondYearAdmission.DoesNotExist:
        existing_app = None

    if request.method == 'POST':
        form = SecondYearAdmissionForm(request.POST)
        if form.is_valid():
            admission = form.save(commit=False)
            admission.user = request.user
            admission.save()
            return redirect('admissionapp:second_year_admission_detail')
    else:
        form = SecondYearAdmissionForm()

    return render(request, 'second_year_form.html', {'form': form})

@login_required
def first_year_admission_detail(request):
    try:
        app = request.user.first_year_application
    except FirstYearAdmission.DoesNotExist:
        return redirect('admissionapp:first_year_admission')

    return render(request, 'first_year_detail.html', {'app': app})

@login_required
def second_year_admission_detail(request):
    try:
        app = request.user.second_year_application
    except SecondYearAdmission.DoesNotExist:
        return redirect('admissionapp:second_year_admission')

    return render(request, 'second_year_detail.html', {'app': app})


def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registration successful. You can now log in.")
            return redirect('admissionapp:login')
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form})

def forgot_password(request):
    return render(request, 'forgot_password.html')



@login_required
def download_application_pdf(request):
    fe_app = getattr(request.user, 'first_year_application', None)
    se_app = getattr(request.user, 'second_year_application', None)
    print("Asad",fe_app, se_app)
    if fe_app:
        is_fe_app = True
        context = {'app': fe_app, 'is_fe_app' : is_fe_app}
        template_name = 'application_pdf.html'
    elif se_app:
        context = {'app': se_app}
        template_name = 'application_pdf.html'
    else:
        # Handle no application case
        messages.error(request, "No application found to download.")
        return redirect('admissionapp:dashboard')
    print(context)
    html_string = render_to_string(template_name, context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="CAP_Application_{}.pdf"'.format(request.user.username)
    return response


@staff_member_required
def first_year_applications_list(request):
    applications = FirstYearAdmission.objects.all().order_by('-submitted_at')
    print(applications)
    return render(request, 'first_year_applications_list.html', {'applications': applications})

@staff_member_required
def second_year_applications_list(request):
    applications = SecondYearAdmission.objects.all().order_by('-submitted_at')
    return render(request, 'second_year_applications_list.html', {'applications': applications})

@user_passes_test(lambda u: u.is_staff)
def first_year_application_detail(request, pk):
    app = get_object_or_404(FirstYearAdmission, pk=pk)
    return render(request, 'first_year_admission_detail.html', {'app': app, 'admin_mode': True})

@staff_member_required
def second_year_application_detail(request, pk):
    app = get_object_or_404(SecondYearAdmission, pk=pk)
    return render(request, 'second_year_admission_detail.html', {'app': app, 'admin_mode': True})

@staff_member_required
def download_first_year_application_pdf(request, pk):
    app = get_object_or_404(FirstYearAdmission, pk=pk)
    html_string = render_to_string('application_pdf.html', {'app': app})
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="FirstYear_Application_{app.id}.pdf"'
    return response

@staff_member_required
def download_second_year_application_pdf(request, pk):
    app = get_object_or_404(SecondYearAdmission, pk=pk)
    html_string = render_to_string('application_pdf.html', {'app': app})
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="SecondYear_Application_{app.id}.pdf"'
    return response

@staff_member_required
def export_applications_excel(request):
    wb = Workbook()

    # ===== First Sheet: First Year Applications =====
    ws1 = wb.active
    ws1.title = "First Year Applications"

    fe_headers = [
        "Full Name", "Religion", "DOB", "Student Cell", "Student Email",
        "Nationality", "MHT CET Percentile", "State Merit No", "Application ID"
    ]
    ws1.append(fe_headers)

    fe_apps = FirstYearAdmission.objects.all().order_by('-submitted_at')
    for app in fe_apps:
        dob = f"{app.dob_day:02d}/{app.dob_month:02d}/{app.dob_year}"
        ws1.append([
            app.full_name,
            app.religion,
            dob,
            app.student_cell_no,
            app.student_email,
            app.nationality,
            getattr(app, 'mht_cet_percentile', ''),
            getattr(app, 'state_merit_no', ''),
            app.application_id,
        ])

    # Adjust column widths for FE
    for i, col in enumerate(ws1.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws1.column_dimensions[chr(64 + i)].width = max_length + 3

    # ===== Second Sheet: Second Year Applications =====
    ws2 = wb.create_sheet(title="Second Year Applications")

    se_headers = [
        "Full Name", "Religion", "DOB", "Student Cell", "Student Email",
        "Nationality", "Passed Diploma Branch", "Diploma Passing Percentage", "Application ID"
    ]
    ws2.append(se_headers)

    se_apps = SecondYearAdmission.objects.all().order_by('-submitted_at')
    for app in se_apps:
        dob = f"{app.dob_day:02d}/{app.dob_month:02d}/{app.dob_year}"
        ws2.append([
            app.full_name,
            app.religion,
            dob,
            app.student_cell_no,
            app.student_email,
            app.nationality,
            getattr(app, 'passed_diploma_branch', ''),
            getattr(app, 'diploma_passing_percentage', ''),
            app.application_id,
        ])

    # Adjust column widths for SE
    for i, col in enumerate(ws2.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws2.column_dimensions[chr(64 + i)].width = max_length + 3

    # Save workbook to in-memory file
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Return response
    response = HttpResponse(
        file_stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=fe_se_applications.xlsx'
    return response
